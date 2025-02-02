import codecs, json, io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from xml.dom.minidom import *

XLSX_FILENAME='categories.xlsx'

#----------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------
class CategoryNode():
    def __init__(self, name, id, local = ''):
        self.name = name
        self.local = local
        self.id = id
        self._id = None
        self.parent_id = None
        self.childs = []
        
    def dump(self, f, deep):
        self.woffset(deep, f)
        f.write(unicode(str(self.name) + '\n', 'utf8'))
        
        if self.childs:
            for child in self.childs:
                child.dump(f, deep + 1)
            
    def woffset(self, deep, f):
        for x in range(0, deep):
            f.write(unicode('  '))

#----------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------
class CategoryTree():
    def __init__(self):
        self.root = None
        
    def find_by_name(self, name):
        '''find category by name'''
        stack = []
        
        stack.append(self.root)
        
        while (len(stack) > 0):
            top = stack.pop(0)
            
            if (top.name == name):
                return top
            
            for child in top.childs:
                stack.insert(0, child) 

        return None

#----------------------------------------------------------------------------------------------
def dump_category_tree(filename, root):
    print("opening damp groups file:" + filename)
    with io.open(filename + '.dump', 'w', encoding='utf8') as f:
        print("opening OK")
        f.write(unicode('{\n'))
        root.dump(f, 0)
        f.write(unicode('}\n'))
        
#----------------------------------------------------------------------------------------------
class CacheData():
    def __init__(self, specs):
        self.specs = specs
        self.cahepath = self.specs['input']['out'] + self.specs['input']['filename'] + '.json'
    
#----------------------------------------------------------------------------------------------    
    def generate(self):
        print("Start caching data...")
        
        fullpath = self.specs['input']['path'] + self.specs['input']['filename']
        
        wb = load_workbook(fullpath)
        
        self.sheet =  wb.active
        row_count = self.sheet.max_row - 1
        max_column = self.sheet.max_column
        
        print('rows:' + str(row_count))
        print('columns:' + str(max_column) + ":" + get_column_letter(max_column))
        
        print("opening cache file:" + self.cahepath)
        with io.open(self.cahepath, 'w', encoding='utf8') as f:
            range = 'A2:' + get_column_letter(max_column) + str(row_count)
            for row in self.sheet.iter_rows(range):
                row_dict = {}
                for cell in row:
                    self.store_cell(cell, row_dict)
                    
                str_row = json.dumps(row_dict, sort_keys=False, ensure_ascii=False).encode('utf8')
                f.write(unicode(str_row + '\n', 'utf8'))
        
        self.generateTreeFromCache()
        
#----------------------------------------------------------------------------------------------    
    def store_cell(self, cell, dict):
        if (cell.value != None):
            if cell.column == 'A':
                dict['Cat0'] = cell.value
            elif cell.column == 'B':
                dict['Cat1'] = cell.value    
            elif cell.column == 'C':
                dict['Cat2'] = cell.value
            elif cell.column == 'D':
                dict['Cat3'] = cell.value
            elif cell.column == 'F':
                dict['id'] = int(cell.value)
                
#----------------------------------------------------------------------------------------------                
    def generateTreeFromCache(self):
        '''load cache and generate tree'''
        
        print("opening cache groups file:" + self.cahepath)
        arr_groups = []
        
        with io.open(self.cahepath , 'r', encoding='utf8') as f:
            for line in f:
                data = json.loads(line)
                arr_groups.append(data)
            
        self.buildTree(arr_groups)
        self.buildXMLTree(arr_groups)
        
        dump_category_tree(self.cahepath + '.tmp', self.tree.root)
        
#----------------------------------------------------------------------------------------------
    def buildTree(self, array):
        ''' build tree from flat structure array'''
        self.tree = CategoryTree()
        self.tree.root = CategoryNode('root', 0)

        for row in array:
            top = self.tree.root
            
            if row.get('Cat0'):
                node = self.tree.find_by_name(row['Cat0'])
                if not node:
                    node = CategoryNode(row['Cat0'], row['id'])
                    top.childs.append(node)
                top = node
                    
            if row.get('Cat1'):
                node = self.tree.find_by_name(row['Cat1'])
                if not node:
                    node = CategoryNode(row['Cat1'], row['id'])
                    top.childs.append(node)
                top = node
                    
            if row.get('Cat2'):
                node = self.tree.find_by_name(row['Cat2'])
                if not node:
                    node = CategoryNode(row['Cat2'], row['id'])
                    top.childs.append(node)
                top = node
                    
            if row.get('Cat3'):
                node = self.tree.find_by_name(row['Cat3'])
                if not node:
                    node = CategoryNode(row['Cat3'], row['id'])
                    top.childs.append(node)
                top = node

#----------------------------------------------------------------------------------------------
    def buildXMLTree(self, array):
        print('buildXMLTree')
        doc = Document()
        root = doc.createElement('node')
        root.setAttribute("name", "root")
        doc.appendChild(root)
        
        for row in array:
            nodes =  []
            if row.get('Cat0'):
                nodes.append(row['Cat0'])
            if row.get('Cat1'):
                nodes.append(row['Cat1'])
            if row.get('Cat2'):
                nodes.append(row['Cat2'])            
            if row.get('Cat3'):
                nodes.append(row['Cat3'])
                
            self.addXML(nodes, row['id'], root, doc)
        
        out = self.specs['input']['out'] + self.specs['opt']['aspect'] + '.xml'
        print('save to {}'.format(out))
        doc.writexml(open(out, 'w'),
                       indent=" ",
                       addindent= "    ",
                       newl='\n')
        
    #----------------------------------------------------------------------------------------------
    def addXML(self, names, foreign_id, root, doc):
        node_to_add = root
        
        n_count = 0
        for name in names:
            bAdded = False
            childs = node_to_add.childNodes
            for child in childs:
                if child.nodeType == child.ELEMENT_NODE and child.localName == 'node':
                    if child.getAttribute("name") == name:
                        node_to_add = child
                        bAdded = True
                        break
                        
            if not bAdded:
                str_foreign_id = ''
                if n_count == len(names) - 1: # add foreign id to last trailing node
                    str_foreign_id = str(foreign_id)
                node_to_add = self.pushNode(node_to_add, name, str_foreign_id , doc)
            
            n_count += 1
            
    #----------------------------------------------------------------------------------------------
    def pushNode(self, node, name, str_foreing_id, doc):
        new_node = doc.createElement('node')
        new_node.setAttribute("name", name)
        new_node.setAttribute("local", "")
        new_node.setAttribute("foreign_id", str_foreing_id)
        node.appendChild(new_node)
        return new_node