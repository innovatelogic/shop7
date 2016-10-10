import os, sys, shutil, argparse
import codecs, json, io
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

class CacheItemsDB:
	def __init__(self, filename, sheet):
		self.filename = filename
		self.sheet = sheet
		
	def generate(self):
		print("generate items cache...")
		
		row_count = self.sheet.max_row - 1
		max_column = self.sheet.max_column
		
		print('rows:' + str(row_count))
		print('columns:' + str(max_column) + ":" + get_column_letter(max_column))
		
		print("opening cache file:" + self.filename)
		with io.open(self.filename, 'w', encoding='utf8') as f:
			range = 'B2:' + get_column_letter(max_column) + str(row_count)
			for row in self.sheet.iter_rows(range):
				row_dict = {}
				for cell in row:
					self.store_cell(cell, row_dict)
				
				str_json = json.dumps(row_dict, sort_keys=False, ensure_ascii=False).encode('utf8')
				f.write(unicode(str_json + '\n', 'utf8'))
		print("items cache OK")
		
	def store_cell(self, cell, dict):
		if (cell.value != None):
			if cell.column == 'B':
				dict['name'] = cell.value
			elif cell.column == 'C':
				dict['keywords'] = cell.value
			elif cell.column == 'D':
				dict['desc'] = cell.value
			elif cell.column == 'E':
				dict['type'] = cell.value
			elif cell.column == 'F':
				dict['price'] = cell.value
			elif cell.column == 'G':
				dict['currency'] = cell.value
			elif cell.column == 'H':
				dict['unit'] = cell.value
			elif cell.column == 'L':
				dict['imageRefs'] = cell.value
			elif cell.column == 'M':
				dict['availability'] = cell.value
			elif cell.column == 'N':
				dict['amount'] = cell.value	
			elif cell.column == 'O':
				dict['groupID'] = cell.value		
			elif cell.column == 'P':
				dict['prom_ua_subsection'] = cell.value		
			elif cell.column == 'T':
				dict['uniqID'] = cell.value			
			elif cell.column == 'U':
				dict['productID'] = cell.value							
			elif cell.column == 'V':
				dict['subsectionID'] = cell.value					
			elif cell.column == 'W':
				dict['groupID2'] = cell.value					
			elif cell.column == 'AC':
				dict['characteristicName0'] = cell.value					
			elif cell.column == 'AE':
				dict['characteristicValue0'] = cell.value				
			elif cell.column == 'AF':
				dict['characteristicName1'] = cell.value				
			elif cell.column == 'AH':
				dict['characteristicValue1'] = cell.value								
			elif cell.column == 'AI':
				dict['characteristicName2'] = cell.value				
			elif cell.column == 'AK':
				dict['characteristicValue2'] = cell.value				
			elif cell.column == 'AL':
				dict['characteristicName3'] = cell.value				
			elif cell.column == 'AN':
				dict['characteristicValue3'] = cell.value					
			elif cell.column == 'AO':
				dict['characteristicName4'] = cell.value				
			elif cell.column == 'AQ':
				dict['characteristicValue4'] = cell.value				
			elif cell.column == 'AR':
				dict['characteristicName5'] = cell.value				
			elif cell.column == 'AT':
				dict['characteristicValue5'] = cell.value					
			elif cell.column == 'AU':
				dict['characteristicName6'] = cell.value				
			elif cell.column == 'AW':
				dict['characteristicValue6'] = cell.value				
				
				
				
				