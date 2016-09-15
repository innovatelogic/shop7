import os, sys, shutil, argparse
import codecs, json, io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

class CacheGroupsDB:
	def __init__(self, filename, sheet):
		self.filename = filename
		self.sheet = sheet
		
	def generate(self):
		print("generate group cache...")
		
		row_count = self.sheet.max_row - 1
		max_column = self.sheet.max_column
		
		print('rows:' + str(row_count))
		print('columns:' + str(max_column) + ":" + get_column_letter(max_column))
		
		print("opening cache file:" + self.filename)
		with io.open(self.filename, 'w', encoding='utf8') as f:
			range = 'A2:' + get_column_letter(max_column) + str(row_count)
			for row in self.sheet.iter_rows(range):
				row_dict = {}
				for cell in row:
					self.store_cell(cell, row_dict)
					
				str_row = json.dumps(row_dict, sort_keys=False, ensure_ascii=False).encode('utf8')

				f.write(unicode(str_row + '\n', 'utf8'))
				
	def store_cell(self, cell, dict):
		if (cell.value != None):
			if cell.column == 'A':
				dict['GroupNumber'] = cell.value
			elif cell.column == 'B':
				dict['GroupName'] = cell.value	
			elif cell.column == 'C':
				dict['GroupID'] = cell.value
			elif cell.column == 'D' and cell.value:
				dict['GroupParentNumber'] = cell.value
			elif cell.column == 'E' and cell.value:
				dict['GroupParentID'] = cell.value			
			
			
			
			
			
				