import os, sys, shutil, argparse
import codecs, json, io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

class CacheItemsDB:
	def __init__(self, filename, sheet):
		self.filename = filename
		self.sheet = sheet
		
	def generate(self):
		print("generate...")
		
		row_count = self.sheet.max_row - 1
		max_column = self.sheet.max_column
		
		print('rows:' + str(row_count))
		print('columns:' + str(max_column) + ":" + get_column_letter(max_column))
		
		print("opening cache file")
		with io.open(self.filename, 'w', encoding='utf8') as f:
			range = 'B1:' + get_column_letter(max_column) + str(row_count)
			for row in self.sheet.iter_rows(range):
				row_dict = {}
				for cell in row:
					self.store_cell(cell, row_dict)
			
				json_str = json.dumps(row_dict, ensure_ascii=False)
				#json.dump(row_dict, f)
				f.write(json_str + '\n')
				
	def store_cell(self, cell, dict):
		if (cell.value != None):
			if cell.column == 'B':
				dict['Name'] = cell.value
			elif cell.column == 'C':
				dict['Keywords'] = cell.value
			elif cell.column == 'D':
				dict['Desc'] = cell.value
			elif cell.column == 'E':
				dict['Type'] = cell.value
			elif cell.column == 'F':
				dict['Price'] = cell.value
			elif cell.column == 'G':
				dict['Currency'] = cell.value
			elif cell.column == 'H':
				dict['Unit'] = cell.value
			elif cell.column == 'L':
				dict['ImageRefs'] = cell.value
			elif cell.column == 'M':
				dict['Availability'] = cell.value
			elif cell.column == 'N':
				dict['Amount'] = cell.value	
			elif cell.column == 'O':
				dict['GroupID'] = cell.value		
			elif cell.column == 'P':
				dict['Prom_ua_subsection'] = cell.value		
			elif cell.column == 'T':
				dict['UniqID'] = cell.value			
			elif cell.column == 'U':
				dict['ProductID'] = cell.value							
			elif cell.column == 'V':
				dict['SubsectionID'] = cell.value					
			elif cell.column == 'W':
				dict['GroupID2'] = cell.value					
			elif cell.column == 'AC':
				dict['characteristicName0'] = cell.value					
			elif cell.column == 'AE':
				dict['characteristicValue0'] = cell.value				
			elif cell.column == 'AF':
				dict['characteristicName1'] = cell.value				
			elif cell.column == 'AH':
				dict['characteristicValue1'] = cell.value								
			elif cell.column == 'AI':
				dict['characteristicName2'] = cell.value				
			elif cell.column == 'AK':
				dict['characteristicValue2'] = cell.value				
			elif cell.column == 'AL':
				dict['characteristicName3'] = cell.value				
			elif cell.column == 'AN':
				dict['characteristicValue3'] = cell.value					
			elif cell.column == 'AO':
				dict['characteristicName4'] = cell.value				
			elif cell.column == 'AQ':
				dict['characteristicValue4'] = cell.value				
			elif cell.column == 'AR':
				dict['characteristicName5'] = cell.value				
			elif cell.column == 'AT':
				dict['characteristicValue5'] = cell.value					
			elif cell.column == 'AU':
				dict['characteristicName6'] = cell.value				
			elif cell.column == 'AW':
				dict['characteristicValue6'] = cell.value				
				
				
				
				