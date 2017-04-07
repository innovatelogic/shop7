import os, sys, shutil, argparse
import codecs, json, io
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from common.db.types.category import Category
from common.aspect import Aspect, CategoryNode
from common.models.base_aspects_container import BaseAspectsContainer, BaseAspectHelper

#----------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------
class CacheGroupsDB:
	def __init__(self, sheet):
		self.sheet = sheet
		root_category = Category({'_id': -1, 'parent_id': None, 'name':'root'})
		root_category_node = CategoryNode(root_category, None)
		self.aspect = Aspect('', root_category_node)
		self.group_key_hash = {}

#----------------------------------------------------------------------------------------------
	def generate(self):
		print("generate group cache...")
		
		row_count = self.sheet.max_row - 1
		max_column = self.sheet.max_column
		
		print('rows:' + str(row_count))
		print('columns:' + str(max_column) + ":" + get_column_letter(max_column))
		
		range = 'A3:' + get_column_letter(max_column) + str(row_count)
		rows = []
		
		for row in self.sheet.iter_rows(range):
			dict = {}
			for cell in row:
				self.store_cell(cell, dict)
			rows.append(dict)
			
		stack = []
		stack.append(self.aspect.root)
		
		while stack:
			top = stack.pop(0)
			
			for row in rows:
				if row and str(top.category._id) == row['GroupParentNumber']:
					node = self.addCategoryImpl(Category({'_id': row['GroupNumber'], 
														  'parent_id': row['GroupParentNumber'],
														  'name':row['GroupName']}), top)
					stack.append(node)
					if 'GroupID' in row:
						self.group_key_hash[str(row['GroupID'])] = node
					row = None
					
		#BaseAspectHelper.dump_category_tree('D:\dump_file.txt', self.aspect.root)

#----------------------------------------------------------------------------------------------
	def store_cell(self, cell, dict):
		if (cell.value != None):
			if cell.column == 'A':
				dict['GroupNumber'] = str(cell.value)
			elif cell.column == 'B':
				dict['GroupName'] = str(cell.value)
			elif cell.column == 'C':
				dict['GroupID'] = str(cell.value)
			elif cell.column == 'D':
				if cell.value:
					dict['GroupParentNumber'] = str(cell.value)
				else:
					dict['GroupParentNumber'] = str(-1)
			elif cell.column == 'E' and cell.value:
					dict['GroupParentID'] = cell.value

#----------------------------------------------------------------------------------------------			
	def addCategoryImpl(self, category, parent_node):
		new_node = CategoryNode(category, parent_node)
		parent_node.childs.append(new_node)
		self.aspect.hashmap[str(category._id)] = new_node
		return new_node